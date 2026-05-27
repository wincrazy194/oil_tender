"""
数据存储模块 - 负责 SQLite 持久化、去重和文件导出
"""
import sqlite3
import pandas as pd
import logging
import os
from datetime import datetime
from config import DB_PATH, EXPORT_DIR

logger = logging.getLogger(__name__)

class Storage:
    def __init__(self):
        self._init_db()

    def _init_db(self):
        """初始化数据库表"""
        with sqlite3.connect(DB_PATH) as conn:
            cursor = conn.cursor()
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS tenders (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    company TEXT,
                    title TEXT,
                    category TEXT,
                    publish_date TEXT,
                    url TEXT UNIQUE,
                    content TEXT,
                    summary TEXT,
                    fetched_at TEXT,
                    is_it_related BOOLEAN DEFAULT 1
                )
            """)
            # 检查是否需要添加 content 列（旧数据库升级）
            try:
                cursor.execute("ALTER TABLE tenders ADD COLUMN content TEXT")
                conn.commit()
                logger.info("已添加 content 列")
            except sqlite3.OperationalError:
                pass  # 列已存在

            # 检查是否需要添加 summary 列
            try:
                cursor.execute("ALTER TABLE tenders ADD COLUMN summary TEXT")
                conn.commit()
                logger.info("已添加 summary 列")
            except sqlite3.OperationalError:
                pass  # 列已存在

            # 检查是否需要添加 is_it_related 列（AI 判断结果）
            try:
                cursor.execute("ALTER TABLE tenders ADD COLUMN is_it_related BOOLEAN DEFAULT 1")
                conn.commit()
                logger.info("已添加 is_it_related 列")
            except sqlite3.OperationalError:
                pass  # 列已存在

            conn.commit()
        logger.info(f"数据库就绪：{DB_PATH}")

    def save_records(self, records: list[dict]) -> int:
        """保存 IT 相关记录（AI 已判断），返回新增条数"""
        new_count = 0
        if not records:
            return 0

        with sqlite3.connect(DB_PATH) as conn:
            cursor = conn.cursor()
            for r in records:
                try:
                    # 只保存 AI 判定为 IT 相关的记录
                    if not r.get("is_it_related", False):
                        continue

                    category = "IT 业务-" + r["category"]
                    content = r.get("content", "")  # 获取详情内容
                    summary = r.get("summary", "")  # 获取概述
                    is_it_related = r.get("is_it_related", True)

                    cursor.execute("""
                        INSERT INTO tenders (company, title, category, publish_date, url, content, summary, fetched_at, is_it_related)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """, (r["company"], r["title"], category, r["publish_date"], r["url"], content, summary, r["fetched_at"], is_it_related))
                    new_count += 1
                except sqlite3.IntegrityError:
                    # 链接重复，忽略
                    continue
            conn.commit()
        return new_count

    def get_todays_records(self) -> list[dict]:
        """获取今天采集的所有 IT 相关记录"""
        today = datetime.now().strftime("%Y-%m-%d")
        with sqlite3.connect(DB_PATH) as conn:
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()
            # 只获取 IT 相关记录（fetched_at 包含今天日期，且 category 包含 IT 业务）
            cursor.execute("SELECT * FROM tenders WHERE fetched_at LIKE ? AND category LIKE ?", (f"{today}%", "%IT 业务%"))
            rows = cursor.fetchall()
            return [dict(row) for row in rows]

    def export_to_excel(self, records=None) -> str:
        """导出当日数据到 Excel（仅 IT 相关，支持 PDF 超链接）"""
        if records is None:
            records = self.get_todays_records()
        if not records:
            return ""

        # 只保留 IT 相关记录（使用 AI 判断结果）
        it_records = []
        for r in records:
            if r.get("is_it_related", False) or r.get("category", "").startswith("IT 业务"):
                it_records.append(r)

        if not it_records:
            logger.info("无 IT 相关记录，跳过 Excel 导出")
            return ""

        records = it_records  # 使用 IT 记录

        # 使用 openpyxl 创建带超链接的 Excel
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
        except ImportError:
            # 降级使用 pandas
            return self._export_to_excel_pandas(records)

        date_str = datetime.now().strftime("%Y%m%d")
        filename = f"tender_report_{date_str}.xlsx"
        filepath = os.path.join(EXPORT_DIR, filename)

        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "招标公告"

        # 表头
        headers = ["序号", "公司", "标题", "类别", "发布日期", "链接", "PDF", "摘要"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        # 数据行
        pdf_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "output", "cnpc_pdf")
        for row_idx, record in enumerate(records, 2):
            ws.cell(row=row_idx, column=1, value=row_idx - 1)
            ws.cell(row=row_idx, column=2, value=record.get("company", "中石油"))
            ws.cell(row=row_idx, column=3, value=record.get("title", "")[:200])
            ws.cell(row=row_idx, column=4, value=record.get("category", "IT 业务"))
            ws.cell(row=row_idx, column=5, value=record.get("publish_date", ""))

            # 链接列 - 添加超链接
            url = record.get("url", "")
            link_cell = ws.cell(row=row_idx, column=6, value="查看")
            if url:
                link_cell.hyperlink = url
                link_cell.font = Font(color="0000FF", underline="single")
                link_cell.style = "Hyperlink"

            ws.cell(row=row_idx, column=7, value=record.get("summary", "")[:200])

        # 调整列宽
        column_widths = [5, 10, 70, 20, 12, 10, 60]
        for i, width in enumerate(column_widths, 1):
            col_letter = get_column_letter(i)
            ws.column_dimensions[col_letter].width = width

        # 保存文件
        os.makedirs(EXPORT_DIR, exist_ok=True)
        wb.save(filepath)
        logger.info(f"数据已导出：{filepath}")
        return filepath

    def _export_to_excel_pandas(self, records: list[dict]) -> str:
        """降级导出方法（不支持超链接）"""
        df = pd.DataFrame(records)
        date_str = datetime.now().strftime("%Y%m%d")
        filename = f"tender_report_{date_str}.xlsx"
        filepath = os.path.join(EXPORT_DIR, filename)
        df.to_excel(filepath, index=False)
        logger.info(f"数据已导出：{filepath}")
        return filepath
